import math
import re
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ── 工厂识别 ──────────────────────────────────────────────────

def identify_factory(am_value):
    """从 AM 列（通途sku配货名称）判断工厂"""
    if not am_value:
        return None

    tokens = str(am_value).strip().split()
    if not tokens:
        return None

    last = tokens[-1]

    # 彩巢：AM 中包含"彩巢"关键字
    if '彩巢' in str(am_value):
        return '彩巢'

    if last.upper() == 'ML':
        return '浙江曼联'
    if last.upper().startswith('NC-'):
        return '杭州创明'
    if last.upper().startswith('SY-'):
        return '上海顺裕'
    if last.upper().startswith('YS-'):
        return '杭州创明'
    # 日期格式 XX.XX 且首 token 非数字 → 宁波利洋
    if re.match(r'^\d{1,2}\.\d{2}$', last) and not re.match(r'^[\d.]+$', tokens[0]):
        return '宁波利洋'
    # 小数格式 XXX.XXXXX → 广州创明
    if re.match(r'^\d+\.\d+$', last):
        return '广州创明'

    return None


# ── 物流商配对 ────────────────────────────────────────────────

LOGISTICS_MAP = {
    '广州创明': '东莞YDH',
    '杭州创明': '上海YDH',
    '上海顺裕': '上海YDH',
    '宁波利洋': 'OCS-SGW',
    '彩巢': '上海YDH',
}


def width_to_cm(w_str, factory):
    """将宽度字符串转为 cm 单位数值"""
    try:
        w = float(w_str)
    except (ValueError, TypeError):
        return 0
    if factory == '浙江曼联':
        return w * 100  # 米 → cm
    elif factory in ('广州创明', '杭州创明', '上海顺裕'):
        return w / 10  # mm → cm
    else:  # 宁波利洋 / 彩巢
        return w  # cm


def can_ship_ocs(widths_cm, total_qty):
    """检查 OCS 能否发货：箱子三边和 ≤ 220cm"""
    if not widths_cm or total_qty <= 0:
        return True
    max_w = max(widths_cm)
    n = total_qty
    # rows × cols 尽量接近正方形
    rows = math.ceil(math.sqrt(n))
    cols = math.ceil(n / rows)
    box_len = max_w + 6
    box_w = cols * 10
    box_h = rows * 10
    return (box_len + box_w + box_h) <= 220


def get_logistics_for_group(factory, widths_cm, total_qty):
    """根据整个订单（组）计算物流商，同订单不拆分"""
    if factory in ('浙江曼联', '彩巢'):
        max_w = max(widths_cm) if widths_cm else 0
        if max_w > 160:  # > 1.6m = > 160cm
            if can_ship_ocs(widths_cm, total_qty):
                return 'OCS'
            return '上海YDH'
        return '上海YDH'
    return LOGISTICS_MAP.get(factory, '')


# ── AM 列拆分 ─────────────────────────────────────────────────

def parse_am(am_value, factory):
    """将 AM 列拆成 5 部分：产品编码 / 产品名称 / 宽度 / 高度 / 厂家编码"""
    if not am_value or not factory:
        return ['', '', '', '', '']

    tokens = str(am_value).strip().split()
    if len(tokens) < 3:
        return [am_value, '', '', '', '']

    fc = tokens[-1]
    h = tokens[-2] if len(tokens) >= 2 else ''
    w = tokens[-3] if len(tokens) >= 3 else ''

    if factory == '浙江曼联':
        pc = tokens[-4] if len(tokens) >= 4 else ''
        pn = ' '.join(tokens[:-4]) if len(tokens) > 4 else ''
    elif factory in ('宁波利洋', '彩巢'):
        pc = tokens[-4] if len(tokens) >= 4 else ''
        pn = ' '.join(tokens[:-4]) if len(tokens) > 4 else ''
    else:  # 广州创明 / 杭州创明 / 上海顺裕
        pc = tokens[0]
        pn = ' '.join(tokens[1:-3]) if len(tokens) > 4 else ''

    return [pc, pn, w, h, fc]


# ── 分包 ──────────────────────────────────────────────────────

def _split_into_packages(group):
    """将一个订单的 SKU 列表按宽度规则拆分为多个包裹"""
    # 复制并展开为独立产品单元
    units = []
    for r in group:
        qty = int(float(r['sku数量'])) if r['sku数量'] else 0
        w = r['width_cm']
        for _ in range(qty):
            units.append({'row': r, 'width': w, 'is_wide': w > 165})

    if not units:
        return [group]  # 无产品，保持原样

    has_wide = any(u['is_wide'] for u in units)
    has_narrow = any(not u['is_wide'] for u in units)

    if has_wide and not has_narrow:
        max_per_pkg = 2
        wide_per_pkg = 2
    elif has_narrow and not has_wide:
        max_per_pkg = 4
        wide_per_pkg = 4
    else:
        max_per_pkg = 4
        wide_per_pkg = 1

    # 排序：宽的在前
    units.sort(key=lambda u: (0 if u['is_wide'] else 1, -u['width']))

    # 贪心装箱
    packages = []
    used = [False] * len(units)

    while not all(used):
        pkg_units = []
        pkg_wide = 0
        for i, u in enumerate(units):
            if used[i]:
                continue
            if len(pkg_units) >= max_per_pkg:
                break
            if u['is_wide'] and pkg_wide >= wide_per_pkg:
                continue

            pkg_units.append(i)
            used[i] = True
            if u['is_wide']:
                pkg_wide += 1

        packages.append(pkg_units)

    # 将包裹转回行格式：按原始 SKU 聚合数量
    result = []
    for pkg in packages:
        # 统计此包裹中每个原始行的产品数
        counts = {}
        for idx in pkg:
            row_id = id(units[idx]['row'])
            counts[row_id] = counts.get(row_id, 0) + 1

        pkg_rows = []
        for row_id, qty in counts.items():
            # 找到原始行
            orig = next(r for r in group if id(r) == row_id)
            new_row = dict(orig)
            new_row['sku数量'] = str(qty)
            pkg_rows.append(new_row)

        result.append(pkg_rows)

    return result


# ── 主处理函数 ─────────────────────────────────────────────────

def process(input_path):
    """读取未整理 xlsx，返回 (output_rows, headers, errors)"""
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active

    # ── 读表头，建立列名 → 列索引映射 ──
    headers = [cell.value for cell in ws[1]]
    col_map = {str(h): i for i, h in enumerate(headers) if h}

    def get_val(row, name, default=''):
        idx = col_map.get(name)
        if idx is None:
            return default
        val = row[idx].value
        return str(val).strip() if val is not None else ''

    # ── 逐行读取为原始记录 ──
    raw_rows = []
    errors = []
    failed_orders = []  # 无法识别的订单号列表

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
        order_no = get_val(row, '订单号')
        am_raw = get_val(row, '通途sku配货名称')

        if not am_raw:
            continue  # 空行跳过

        factory = identify_factory(am_raw)
        if factory is None:
            errors.append(f'第 {row_num} 行：无法识别工厂，AM列内容 = "{am_raw}"')
            failed_orders.append({'order_no': order_no, 'row': row_num, 'am': am_raw})
            continue

        pc, pn, w, h, fc = parse_am(am_raw, factory)
        width_cm = width_to_cm(w, factory)

        # 组装城市：省/州 + 收货地址1 + 收货地址2
        city_parts = [
            get_val(row, '省/州'),
            get_val(row, '收货地址1'),
            get_val(row, '收货地址2'),
        ]
        city = ''.join(p for p in city_parts if p)

        raw_rows.append({
            'order_no': order_no,
            '生成日期': get_val(row, '生成日期'),
            '买家姓名': get_val(row, '买家姓名'),
            '联系电话': get_val(row, '联系电话'),
            '邮编': get_val(row, '邮编'),
            '城市': city,
            '通途sku': get_val(row, '通途sku'),
            '产品编码': pc,
            '产品名称': pn,
            '宽度': w,
            '高度': h,
            '厂家编码': fc,
            'sku数量': get_val(row, '通途sku数量'),
            '通途sku货品备注': get_val(row, '通途sku货品备注'),
            '平台sku': get_val(row, '平台sku'),
            '平台sku数量': get_val(row, '平台sku数量'),
            '订单备注': get_val(row, '订单备注'),
            '买家留言': get_val(row, '买家留言'),
            'factory': factory,
            'width_cm': width_cm,
        })

    # ── 同订单号归组 ──
    groups = []  # (order_no, [raw_rows...])
    i = 0
    while i < len(raw_rows):
        group = [raw_rows[i]]
        current_order = raw_rows[i]['order_no']
        j = i + 1
        while j < len(raw_rows):
            if raw_rows[j]['order_no'] and raw_rows[j]['order_no'] != current_order:
                break
            if not raw_rows[j]['order_no'] or raw_rows[j]['order_no'] == current_order:
                group.append(raw_rows[j])
                j += 1
            else:
                break

        # 计算组级别的物流商（同订单不拆分）
        total_qty = 0
        widths_cm = []
        for r in group:
            try:
                total_qty += int(float(r['sku数量']))
            except (ValueError, TypeError):
                pass
            widths_cm.append(r['width_cm'])

        factory = group[0]['factory']
        logistics = get_logistics_for_group(factory, widths_cm, total_qty)
        logistics_factory = f'{factory}-{logistics}'

        # 统一设置物流商
        for r in group:
            r['物流商与厂家'] = logistics_factory

        groups.append((current_order, group))
        i = j

    # ── 分包：按宽度规则将订单拆分为包裹 ──
    pkg_groups = []  # (order_no, [rows_for_this_package])
    for order_no, group in groups:
        packages = _split_into_packages(group)
        base_order = order_no
        for pkg_idx, pkg_rows in enumerate(packages):
            if pkg_idx == 0:
                pkg_order = base_order
            else:
                pkg_order = f'{base_order}-{pkg_idx}'
            # 更新每行的订单号
            for r in pkg_rows:
                r['order_no'] = pkg_order
            pkg_groups.append((pkg_order, pkg_rows))

    # ── 按厂家排序，同厂家保持原始顺序 ──
    FACTORY_ORDER = {
        '广州创明-东莞YDH': 0,
        '杭州创明-上海YDH': 1,
        '上海顺裕-上海YDH': 2,
        '宁波利洋-OCS-SGW': 3,
        '彩巢-上海YDH': 4,
        '彩巢-OCS': 5,
        '浙江曼联-上海YDH': 6,
        '浙江曼联-OCS': 7,
    }
    pkg_groups.sort(key=lambda g: FACTORY_ORDER.get(g[1][0]['物流商与厂家'], 99))

    # ── 按组生成输出行，同时记录需要合并的单元格范围 ──
    output_rows = []
    merge_ranges = []  # (start_row, end_row, start_col, end_col) 1-indexed

    for order_no, group in pkg_groups:
        # 计算订单总 SKU 数量（已在上面的循环中计算，此处重算保证一致性）
        total_qty = 0
        for r in group:
            try:
                total_qty += int(float(r['sku数量']))
            except (ValueError, TypeError):
                pass

        start_row = len(output_rows) + 2  # +2 因为 row 1 是表头

        for idx, r in enumerate(group):
            # 浙江曼联：J列(产品编码)与K列(产品名称)互换
            pc_val = r['产品名称'] if '浙江曼联' in r['物流商与厂家'] else r['产品编码']
            pn_val = r['产品编码'] if '浙江曼联' in r['物流商与厂家'] else r['产品名称']

            if idx == 0:
                output_rows.append([
                    r['物流商与厂家'],
                    r['生成日期'],
                    r['order_no'],
                    r['买家姓名'],
                    r['联系电话'],
                    r['邮编'],
                    r['城市'],
                    str(total_qty),
                    r['通途sku'],
                    pc_val,
                    pn_val,
                    r['宽度'],
                    r['高度'],
                    r['厂家编码'],
                    r['sku数量'],
                    r['通途sku货品备注'],
                    r['平台sku'],
                    r['平台sku数量'],
                    r['订单备注'],
                    r['买家留言'],
                ])
            else:
                output_rows.append([
                    '', '', '', '', '', '', '', '',
                    r['通途sku'],
                    pc_val,
                    pn_val,
                    r['宽度'],
                    r['高度'],
                    r['厂家编码'],
                    r['sku数量'],
                    r['通途sku货品备注'],
                    r['平台sku'],
                    r['平台sku数量'],
                    r['订单备注'],
                    r['买家留言'],
                ])

        # 多 SKU 订单（group > 1 行）：前 8 列需要合并
        if len(group) > 1:
            end_row = start_row + len(group) - 1
            for col in range(1, 9):  # 列 1-8：物流商与厂家 → 通途sku数量
                merge_ranges.append((start_row, end_row, col, col))

    # ── 输出表头 ──
    out_headers = [
        '物流商与厂家', '生成日期', '订单号', '买家姓名', '联系电话',
        '邮编', '城市', '包裹内产品数', '通途sku', '通途sku配货名称',
        '产品名称', '宽度', '高度', '厂家编码', '通途sku数量',
        '通途sku货品备注', '平台sku', '平台sku数量', '订单备注', '买家留言',
    ]

    return output_rows, out_headers, errors, merge_ranges, failed_orders


# 需要以文本格式存储的列（防止 Excel 自动转数字丢失尾零）
TEXT_FORMAT_COLS = {10, 12, 13, 14}  # 产品编码、宽度、高度、厂家编码


def write_output(output_rows, headers, output_path, merge_ranges=None, failed_orders=None):
    """将处理结果写入 xlsx 文件，可合并单元格"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet0'
    ws.freeze_panes = 'A2'  # 冻结首行

    # 写表头
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)

    # 设置列宽：A(物流商与厂家)=22, B(生成日期)=14, C(订单号)=30
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 30

    # 写数据
    for row_idx, row_data in enumerate(output_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else '')
            if col_idx in TEXT_FORMAT_COLS and val:
                cell.number_format = '@'

    # 合并单元格
    if merge_ranges:
        for start_row, end_row, start_col, end_col in merge_ranges:
            ws.merge_cells(
                start_row=start_row, start_column=start_col,
                end_row=end_row, end_column=end_col,
            )
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    ws.cell(row=r, column=c).alignment = Alignment(vertical='center')

    # 追加未能处理的订单号（红字）
    if failed_orders:
        last_row = len(output_rows) + 2  # 数据末尾下一行
        ws.cell(row=last_row, column=1, value='以下订单未能识别工厂，请检查：').font = Font(color='FF0000')
        red_font = Font(color='FF0000')
        for i, fo in enumerate(failed_orders):
            r = last_row + 1 + i
            cell = ws.cell(row=r, column=3, value=fo['order_no'])
            cell.font = red_font
            ws.cell(row=r, column=10, value=fo['am']).font = red_font

    wb.save(output_path)
